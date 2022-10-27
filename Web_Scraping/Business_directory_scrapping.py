# -*- coding: utf-8 -*-
"""
Created on Fri Mar 12 17:41:47 2021

@author: Camille Souvigny
"""

from fake_useragent import UserAgent
from bs4 import BeautifulSoup
import requests
import urllib3
from pandas import DataFrame, ExcelWriter
from random import randint
from time import sleep
from openpyxl import load_workbook
from datetime import datetime, time
import os

# Excel File 
path = r"C:/Users/camom/OneDrive/Documents/perso"
sleep_min = 7
sleep_max = 13

# user_agent différent selon l'ordi uttilisée et selon on fait un accès par chrome ou mozzila
ua = UserAgent()
user_agent=ua.chrome

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)    # Pour ne pas avoir l'erreur bad handshake



def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, **to_excel_kwargs):

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = ExcelWriter(filename, engine='openpyxl')
    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()

 


def get_link_annuaire(session):
    a_enlever=['/rejoindre-notre-communaute'
    '/conditions-generales-de-vente',
    '/statuts',
    '/mentions-legales',
    '/politique-de-confidentialite',
    '/formulaire-de-contact'
    '/histoire-de-lafte',
    '/lafte-et-les-etudiants',
    '/emplois-et-stages',
    '/adherer-lafte',
    '/mes-favoris',
    '/histoire-de-lafte',
    '/lafte-et-les-etudiants',
    '/recherche-formation',
    '/evenements-partenaires',
    '/emplois-et-stages',
    '/rejoindre-notre-communaute',
    '/conditions-generales-de-vente',
    '/mentions-legales',
    '/politique-de-confidentialite',
    '/formulaire-de-contact']
    
    
    annuaire_page='https://www.XXX.com/annuaire?nom=&field_region_target_id=All&societe=&page='
    lien_annuaire=[] #On y stocke les chemins relatifs d'accès aux page annuaire individuelles
   
    for i in range(141):
        
        annuaire=annuaire_page+str(i) #pages intermédiaires avec le lien des page annuaires
        sleep(randint(10, 20)) # Si moins de temps par requête le AFTE bloque l'accès
        
        html_content= session.get(annuaire, headers={"User-Agent": user_agent}, verify=False)
        soup = BeautifulSoup(html_content.text, "html.parser")
        
        
        for link in soup.find_all("a"):  # get all links
            if link.get("href")!=None:
                if "-" in link.get("href") and link.get("href")[0]=='/' and link.get("href").count('/')==1 and link.get("href") not in a_enlever:
                    lien_annuaire.append(link.get("href"))
      
    return lien_annuaire

 


def get_info_annuaire(session,url,nom):

    sleep(randint(sleep_min, sleep_max))
    
    try:
        html_content = session.get(url, headers={"User-Agent": user_agent}, verify=False)

    # Si le serveur renvoie une erreur, on attend plus longtemps et on retente une fois
    # Si après la deuxième tentative cela ne marche toujours pas, il est plus prudent d'attendre un peu et de relancer le programme
    except:
        sleep(randint(10, 20))
        html_content = session.get(url, headers={"User-Agent": user_agent}, verify=False)
        pass
 
    soup = BeautifulSoup(html_content.text, "html.parser")
    
    #on recupère les informations à partir des balises parsées : aller voir dans le code source les balises
    try:
        #adresse trouvable avec la balise span et le nom de class "adresse-line1
        Adresse=soup.find("span", attrs={"class": "address-line1"}).getText()
    except:
        Adresse='NA'
        pass
    try:
        Code_Postal=soup.find("span", attrs={"class": "postal-code"}).getText()
    except:
        Code_Postal='NA'
        pass
    try:
        Ville=soup.find("span", attrs={"class": "locality"}).getText()
    except:
        Ville='NA'
        pass
    try:
        #société trouvable en cherchant la balise de 'Société' puis en donnant cherchant le contenu de la balise suivante 'div' 
        #On sait que le contenu est dans une balise 'div' 
        Societe=soup.find(text="Société").findNext("div").contents[0]
    except:
        Societe='NA'
        pass
    try:
        Titre=soup.find(text="Titre").findNext("div").contents[0]
    except:
        Titre='NA'
        pass
    try:
        Telephone=soup.find(text="Téléphone").findNext("div").contents[0]
    except:
        Telephone='NA'
        pass
    try:
        Mail=soup.find(text="Mail").findNext("a").contents[0]
    except:
        Mail='NA'
        pass
    try:
        Pays=soup.find("span", attrs={"class": "country"}).getText()
    except:
        Pays='NA'
        pass
    
    if nom[len(nom)-2:len(nom)]=='-0':
            nom=nom[0:len(nom)-2]
    Nom_complet=nom[1:len(nom)]
    
    try:
        Prenom='NA'
        Nom='NA'
        if nom.count('-')==1 :
            pos=nom.find('-')
            Prenom=nom[1:pos]
            Nom=nom[pos+1:len(nom)]
            
        if nom.count('-')==2:
            d=nom.find('-')
            f=nom.find('-',d+1)
 
            if nom[d+1:f]=='de' or nom[d+1:f]=='d' or nom[d+1:f]=='le' or nom[d+1:f]=='da' or nom[d+1:f]=='di' or nom[d+1:f]=='el':
                Prenom=nom[1:d]
                Nom=nom[d+1:f]+" "+ nom[f+1:len(nom)]
                
            elif nom[1:d]=='jean' or nom[1:d]=='anne' or nom[1:d]=='jean' or nom[1:d]=='jean' or nom[1:d]=='paul' or nom[1:d]=='paule' or nom[1:d]=='pierre':
                Prenom=nom[1:f]
                Nom=nom[f+1:len(nom)]
                
                if nom=='/pierre-van-dam':
                    Prenom='pierre'
                    Nom='van dam'
            else:
                Prenom=nom[1:d]
                Nom=nom[d+1:f]+" "+ nom[f+1:len(nom)]

        if nom.count('-')==3:
            d=nom.find('-')
            m=nom.find('-',d+1)
            f=nom.find('-',m+1)
            if nom[d+1:f]=='de-la' or nom[d+1:m]=='d' or nom[d+1:m]=='de' or nom[d+1:m]=='de' or nom[d+1:m]=='d' or nom[d+1:m]=='le' or nom[d+1:m]=='da' or nom[d+1:m]=='di' or nom[d+1:m]=='el' or nom[d+1:m]=='dos':
                Prenom=nom[1:d]
                Nom=nom[d+1:m]+" "+ nom[m+1:f]+" "+ nom[f+1:len(nom)]
                
            elif nom[m+1:f]=='d' or nom[m+1:f]=='de' or nom[m+1:f]=='de' or nom[m+1:f]=='d' or nom[m+1:f]=='le' or nom[m+1:f]=='da' or nom[m+1:f]=='di' or nom[m+1:f]=='el' or nom[m+1:f]=='dos':   
               Prenom=nom[1:d]+" "+nom[d+1:m]
               Nom=nom[m+1:f]+" "+ nom[f+1:len(nom)]
            
    except:
        Prenom='NA'
        Nom='NA'
        pass
                
    return Societe, Titre, Code_Postal, Ville, Adresse, Telephone, Mail, Pays, Prenom, Nom, Nom_complet
 
    
 

def export_to_excel(data, path):

    try:
        if os.path.exists(path + "/" + "Annuaire_AFTE_2021.xlsx"):
            append_df_to_excel(path + "/" + "Annuaire_AFTE_2021.xlsx", data, header=None)
        else:
            data.to_excel(path + "/" + "Annuaire_AFTE_2021.xlsx")
    except PermissionError:
        pass

 
def main():
    titles = ["Société", "Prénom", "Nom", "Nom_complet", "Titre", "Code_Postal", "Ville", "Adresse", "Pays","Téléphone", "Mail", "Lien"]
    data = DataFrame(columns=titles)
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)    
    
    # Website link 
    url_con='https://www.XXX.com/user/login?destination=/'
    
    #Chaque site a des variables de connections différentes : pour les trouver faire F12 sur le site avant de se connecter et recupérer les variables des la requêtes dans un dictionnaire
    login_data = {"name":"X@Y-groupe.com", "pass":"XXX","form_build_id":"form-TE-2BfW3queZ8MkB4cKGb7V_-Q47AirlBj_ORUoYPoQ", "form_id":"user_login_form" ,"op":"VALIDER"}
    # Create session
    session = requests.Session()
    session.post(url_con, login_data)    
    
    # Get directory link
    df_link_dep = get_link_annuaire(session)
    
    http='https://www.XXX.com'
  
    for i in range(0, len(df_link_dep)):  
 
        link=http+df_link_dep[i] #chemin absolu 
        nom=df_link_dep[i]
        
        #recupération des infos dans chaque page de l'annuaire
        Societe, Titre, Code_Postal, Ville, Adresse, Telephone, Mail, Pays, Prenom, Nom, Nom_complet=get_info_annuaire(session,link,nom)
        
        infos_page_annuaire = [Societe, Prenom, Nom, Nom_complet, Titre, Code_Postal, Ville, Adresse, Pays, Telephone, Mail, link]
        data.loc[len(data)] = infos_page_annuaire
        
    # Send frame to Excel 
    data.to_excel(path + "/" + "Scrapping_Result.xlsx")


if __name__ == "__main__":

    main()
